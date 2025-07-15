import re
import threading
import time
import logging
from logging.handlers import RotatingFileHandler
import traceback
import hashlib
import shutil
from pathlib import Path
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import pdfplumber
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from filelock import FileLock
from concurrent.futures import ThreadPoolExecutor
from typing import Dict, Any, Optional
from datetime import datetime, timedelta
from queue import Queue
from threading import Event
import queue
import tkinter as tk
from tkinter import messagebox, simpledialog

from config import config, FRENCH_MONTHS, EXCEL_HEADERS, COUPON_RULES

# Configure logging with rotation
log_filename = f"pv_autofill_{datetime.now().strftime('%Y%m%d')}.log"
log_file = Path(config.paths["LOGS_FOLDER"]) / log_filename

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        RotatingFileHandler(
            str(log_file),  # Convert Path to string for RotatingFileHandler
            maxBytes=5 * 1024 * 1024,  # 5MB
            backupCount=5,
        ),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)

# Data validation patterns
VALIDATION_PATTERNS = {
    "DATE": r"^\d{2}/\d{2}/\d{4}$",
    "N° PV": r"^[A-Z0-9-]+$",
    "IMMATRI": r"^[A-Z0-9 ]+$",  # Updated to allow spaces
    "pht": r"^\d+$",
    "PTTC": r"^\d+$",
    "TVA": r"^\d+$",
}


class ProcessingError(Exception):
    """Custom exception for processing errors."""

    pass


def extract_data(pdf_path):
    """
    Ouvre le PDF, extrait tout le texte et extrait les champs par recherche directe.
    """
    data = {}
    text = ""  # Initialize text variable outside try block
    pdf = None
    try:
        pdf = pdfplumber.open(pdf_path)
        if not pdf.pages:
            raise ValueError("Empty PDF")

        # Extract text from all pages
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)

        # Truncate text at "info@gta-Nomayos.cm" if present
        if "info@gta-Nomayos.cm" in text:
            text = text.split("info@gta-Nomayos.cm")[0]

        lines = text.split("\n")
        data = {key: "" for key in EXCEL_HEADERS}

        # ACCEPTE/REFUS status
        is_rejected = False
        if "ACCEPTE / ACCEPTED" in text:
            data["ACCEPTE"] = "A"
            data["REFUS"] = "-"
            data["C/CV"] = "C"
        elif "CONTRE-VISITE" in text or "REFUSE / REJECTED" in text:
            data["ACCEPTE"] = "-"
            data["REFUS"] = "R"
            data["C/CV"] = "CV"
            is_rejected = True

        # N° PV
        for line in lines:
            if "N° PV" in line and "SHEET" in line:
                parts = line.split(":")
                if len(parts) > 1:
                    data["N° PV"] = parts[1].split("Type")[0].strip()
                    break

        # DATE
        for line in lines:
            if "Date du contrôle" in line and "Date of control" in line:
                parts = line.split(":")
                if len(parts) > 1:
                    date_part = parts[1].split("Catégorie")[0].strip()
                    date_match = re.search(r"(\d{2}/\d{2}/\d{4})", date_part)
                    if date_match:
                        data["DATE"] = date_match.group(1)
                    break

        # DATE P.V - Handle differently for rejected documents
        if is_rejected:
            # For rejected documents, look for "JUSQU'AU / UNTIL" date
            for line in lines:
                if "JUSQU'AU" in line and "UNTIL" in line:
                    parts = line.split(":")
                    if len(parts) > 1:
                        date_part = parts[1].strip()
                        date_match = re.search(r"(\d{2}/\d{2}/\d{4})", date_part)
                        if date_match:
                            data["DATE P.V"] = date_match.group(1)
                            logger.info(
                                f"Found resubmission deadline date: {data['DATE P.V']}"
                            )
                            break
        else:
            # For accepted documents, use "PROCHAINE VISITE"
            for line in lines:
                if "PROCHAINE VISITE" in line and "NEXT VISIT" in line:
                    parts = line.split(":")
                    if len(parts) > 1:
                        date_part = parts[1].split("Type")[0].strip()
                        date_match = re.search(r"(\d{2}/\d{2}/\d{4})", date_part)
                        if date_match:
                            data["DATE P.V"] = date_match.group(1)
                        break

        # IMMATRI
        for line in lines:
            if "Immatriculation" in line and "reg" in line:
                parts = line.split(":")
                if len(parts) > 1:
                    data["IMMATRI"] = parts[1].split("Client")[0].strip()
                    break

        # DESCRIPTIONS (Client info)
        for line in lines:
            if "Client" in line and "Customer" in line:
                parts = line.split(":")
                if len(parts) > 1:
                    data["DESCRIPTIONS"] = parts[2].strip()
                    break

        # CAT (Category field only)
        for line in lines:
            if "Catégorie" in line and "Category" in line:
                parts = line.split(":")
                if len(parts) > 1:
                    string_to_strip = parts[3].strip()
                    data["CAT"] = string_to_strip.split(" ")[0]
                    break

        # CONTACT (Téléphone)
        for line in lines:
            if "Téléphone" in line and "Phone" in line:
                parts = line.split(":")
                if len(parts) > 1:
                    data["CONTACT"] = parts[2].strip()
                    break

        # pht (Montant HT)
        for line in lines:
            if "Montant payé HT" in line:
                parts = line.split(":")
                if len(parts) > 1:
                    amount = parts[2].replace("FCFA", "").replace(" ", "").strip()
                    data["pht"] = amount
                    break

        # PTTC (Montant TTC)
        for line in lines:
            if "Montant payé TTC" in line:
                parts = line.split(":")
                if len(parts) > 1:
                    amount = parts[2].replace("FCFA", "").replace(" ", "").strip()
                    data["PTTC"] = amount
                    break

        # Calculate TVA
        try:
            if data["pht"] and data["PTTC"]:
                data["TVA"] = str(int(data["PTTC"]) - int(data["pht"]))
        except (ValueError, TypeError) as e:
            logger.warning(f"Could not calculate TVA: {str(e)}")
            data["TVA"] = ""

        # Log the extracted values
        for key, value in data.items():
            logger.info(f"Extracted [{key}]: {value}")

        return data

    except Exception as e:
        logger.error(f"Error extracting data from {pdf_path}: {str(e)}")
        logger.debug(traceback.format_exc())
        raise ProcessingError(f"Failed to extract data: {str(e)}")
    finally:
        if pdf:
            pdf.close()


def get_monthly_excel():
    """
    Retourne le workbook du mois courant, en créant le fichier avec entêtes si nécessaire.
    """
    file_path = None
    backup_path = None
    wb = None

    try:
        current_date = datetime.now()
        year = current_date.year
        month = current_date.month
        month_name = FRENCH_MONTHS[month]
        folder_name = f"{month_name} {year}"
        excel_dir = Path(config.paths["EXCEL_FOLDER"]) / str(year) / folder_name
        excel_dir.mkdir(parents=True, exist_ok=True)
        file_name = f"LISTING MOIS {month_name} {year}.xlsx"
        file_path = excel_dir / file_name

        if not file_path.exists():
            # Create new Excel file with headers
            try:
                wb = Workbook()
                ws = wb.active
                ws.append(EXCEL_HEADERS)

                # Style header row - yellow fill but NOT bold
                yellow_fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                )
                # Aptos Narrow font size 11
                aptos_font = Font(
                    name="Aptos Narrow", size=11, bold=False, color="000000"
                )
                left_alignment = Alignment(horizontal="left", wrap_text=False)

                # Apply styles to header row
                for cell in ws[1]:
                    cell.fill = yellow_fill
                    cell.font = aptos_font
                    cell.alignment = left_alignment

                # Auto-adjust column widths to fit content
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        if cell.value:
                            # Calculate the width based on content length
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length

                    # Set width to best fit the content (with some padding)
                    adjusted_width = max_length + 2
                    ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(str(file_path))
                logger.info(f"Created new monthly Excel file: {file_path}")
            except Exception as e:
                logger.error(f"Error creating monthly Excel: {str(e)}")
                # Create a basic workbook as fallback
                wb = Workbook()
                ws = wb.active
                ws.append(EXCEL_HEADERS)
                wb.save(str(file_path))
        else:
            try:
                wb = load_workbook(str(file_path))
                logger.info(f"Opened existing monthly Excel: {file_path}")
            except Exception as e:
                logger.error(f"Error opening existing Excel file: {str(e)}")
                raise

        # Create backup after successfully opening/creating the file
        try:
            backup_path = file_path.with_suffix('.xlsx.bak')
            shutil.copy2(str(file_path), str(backup_path))
            logger.info(f"Created backup at: {backup_path}")
        except Exception as e:
            logger.error(f"Error creating backup: {str(e)}")
            # If backup fails, we should not proceed
            raise

        return str(file_path), wb

    except Exception as e:
        logger.error(f"Error in get_monthly_excel: {str(e)}")
        # Clean up any created files in case of error
        try:
            if backup_path and backup_path.exists():
                backup_path.unlink()
                logger.info(f"Cleaned up backup file after error: {backup_path}")
        except Exception as cleanup_error:
            logger.error(f"Error cleaning up backup file: {str(cleanup_error)}")

        # Create a temporary file as absolute fallback
        try:
            temp_file = Path(config.paths["EXCEL_FOLDER"]) / f"temp_{int(time.time())}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(EXCEL_HEADERS)
            wb.save(str(temp_file))
            logger.info(f"Created temporary Excel file: {temp_file}")
            return str(temp_file), wb
        except Exception as fallback_error:
            logger.error(f"Error creating fallback file: {str(fallback_error)}")
            raise ProcessingError("Could not create or open Excel file")


def append_data(data: Dict[str, Any], wb) -> None:
    """Ajoute une nouvelle ligne dans le workbook ouvert avec les données extraites."""
    try:
        ws = wb.active

        # Validate data before appending
        if not isinstance(data, dict):
            raise ValueError("Data must be a dictionary")

        # Ensure all required headers are present
        missing_headers = [header for header in EXCEL_HEADERS if header not in data]
        if missing_headers:
            logger.warning(f"Missing data for headers: {missing_headers}")
            # Fill missing data with empty strings
            for header in missing_headers:
                data[header] = ""

        # Prepare row data in correct order, converting numeric fields to int if possible
        numeric_fields = ["COUPON", "pht", "TVA", "PTTC"]
        date_fields = ["DATE", "DATE P.V"]
        row = []
        for header in EXCEL_HEADERS:
            value = data.get(header, "")
            if header in numeric_fields and value != "":
                try:
                    value = int(value)
                except (ValueError, TypeError):
                    value = ""
            elif header in date_fields and value:
                try:
                    # Parse DD/MM/YYYY to datetime object
                    value = datetime.strptime(value, "%d/%m/%Y")
                except Exception:
                    pass  # Leave as string if parsing fails
            row.append(value)

        # Add the row
        ws.append(row)

        # Apply cell alignment and font to the newly added row
        left_alignment = Alignment(horizontal="left", wrap_text=False)
        aptos_font = Font(name="Aptos Narrow", size=11, bold=False, color="000000")

        # Get the column headers and their indices
        header_row = ws[1]
        header_indices = {cell.value: idx for idx, cell in enumerate(header_row)}

        # Process the newly added row
        for cell in ws[ws.max_row]:
            cell.alignment = left_alignment
            cell.font = aptos_font
            
            # Get the header for this column
            header = header_row[cell.column - 1].value
            
            # Apply formatting based on the header type
            if header in date_fields and isinstance(cell.value, datetime):
                cell.number_format = "DD/MM/YYYY"
            elif cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.strip()

        # Recalculate column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        continue

            adjusted_width = min(max(max_length + 2, 10), 50)  # Min 10, Max 50
            ws.column_dimensions[column_letter].width = adjusted_width

        logger.info("Data appended successfully to worksheet")
    except Exception as e:
        logger.error(f"Error appending data to Excel: {str(e)}")
        logger.debug(traceback.format_exc())
        raise


def get_file_hash(file_path):
    """
    Calculate SHA-256 hash of a file to uniquely identify it.
    """
    try:
        sha256_hash = hashlib.sha256()
        with open(file_path, "rb") as f:
            # Read file in chunks to handle large files efficiently
            for chunk in iter(lambda: f.read(4096), b""):
                sha256_hash.update(chunk)
        return sha256_hash.hexdigest()
    except Exception as e:
        logger.error(f"Error calculating SHA-256 hash: {str(e)}")
        return "hash-error-" + str(int(time.time()))


def validate_extracted_data(data: Dict[str, Any]) -> None:
    """Validate extracted data against defined patterns and business rules."""
    # Check required fields
    required_fields = ["DATE", "N° PV", "IMMATRI"]
    missing_fields = [field for field in required_fields if not data.get(field)]
    if missing_fields:
        raise ProcessingError(f"Missing required fields: {', '.join(missing_fields)}")

    # Validate field formats
    for field, pattern in VALIDATION_PATTERNS.items():
        if data.get(field) and not re.match(pattern, str(data[field])):
            raise ProcessingError(f"Invalid format for {field}: {data[field]}")

    # Business logic validations
    try:
        # Validate dates
        for date_field in ["DATE", "DATE P.V"]:
            if data.get(date_field):
                datetime.strptime(data[date_field], "%d/%m/%Y")

        # Validate monetary values
        if data.get("pht") and data.get("PTTC"):
            pht = int(data["pht"])
            pttc = int(data["PTTC"])
            if pht >= pttc:
                raise ProcessingError("PTTC must be greater than PHT")
            if pht < 0 or pttc < 0:
                raise ProcessingError("Monetary values cannot be negative")

    except ValueError as e:
        raise ProcessingError(f"Data validation error: {str(e)}")


def extract_data_with_retry(pdf_path: str) -> Dict[str, Any]:
    """Extract data from PDF with retry mechanism."""
    last_exception = None

    for attempt in range(config.settings["RETRY_ATTEMPTS"]):
        try:
            data = extract_data(pdf_path)
            validate_extracted_data(data)
            return data
        except Exception as e:
            last_exception = e
            logger.warning(f"Attempt {attempt + 1} failed: {str(e)}")
            if attempt < config.settings["RETRY_ATTEMPTS"] - 1:
                time.sleep(config.settings["RETRY_DELAY"])

    raise ProcessingError(
        f"Failed after {config.settings['RETRY_ATTEMPTS']} attempts: {str(last_exception)}"
    )


def cleanup_files(excel_path: str) -> None:
    """Clean up temporary files after successful processing."""
    try:
        excel_path_obj = Path(excel_path)
        # Remove lock file if it exists
        lock_file = excel_path_obj.with_suffix('.xlsx.lock')
        if lock_file.exists():
            lock_file.unlink()
            logger.info(f"Removed lock file: {lock_file}")

        # Remove backup file if it exists
        backup_file = excel_path_obj.with_suffix('.xlsx.bak')
        if backup_file.exists():
            backup_file.unlink()
            logger.info(f"Removed backup file: {backup_file}")

    except Exception as e:
        logger.warning(f"Error during cleanup: {str(e)}")


def check_duplicate_file(
    new_file_path: str, target_dir: Path
) -> tuple[bool, Optional[str], Optional[str]]:
    """
    Check if a file is a duplicate by comparing content hashes within the same day's folder.
    Returns (is_duplicate: bool, duplicate_path: Optional[str], pv_number: Optional[str]).
    """
    try:
        # First extract PV number from the new file to correlate with duplicates
        new_file_pv = None
        try:
            with pdfplumber.open(new_file_path) as pdf:
                text = pdf.pages[0].extract_text() or ""
                for line in text.split("\n"):
                    if "N° PV" in line and "SHEET" in line:
                        parts = line.split(":")
                        if len(parts) > 1:
                            new_file_pv = parts[1].split("Type")[0].strip()
                            break
        except Exception as e:
            logger.error(f"Error extracting PV number for duplicate check: {str(e)}")
            return False, None, None

        # Calculate hash of new file
        new_file_hash = get_file_hash(new_file_path)

        # Only check files in the target directory (same day's folder)
        for existing_file in target_dir.glob("*.pdf"):
            if str(existing_file) != new_file_path:  # Don't compare with self
                existing_hash = get_file_hash(str(existing_file))
                if existing_hash == new_file_hash:
                    # If we find a duplicate, extract its PV number for correlation
                    existing_pv = None
                    try:
                        with pdfplumber.open(existing_file) as pdf:
                            text = pdf.pages[0].extract_text() or ""
                            for line in text.split("\n"):
                                if "N° PV" in line and "SHEET" in line:
                                    parts = line.split(":")
                                    if len(parts) > 1:
                                        existing_pv = parts[1].split("Type")[0].strip()
                                        break
                    except Exception:
                        pass

                    logger.warning(
                        f"Duplicate file detected!\n"
                        f"New file: {new_file_path}\n"
                        f"  - PV Number: {new_file_pv or 'Unknown'}\n"
                        f"Existing file: {existing_file}\n"
                        f"  - PV Number: {existing_pv or 'Unknown'}\n"
                        f"{'WARNING: Different PV numbers in duplicate files!' if new_file_pv != existing_pv and new_file_pv and existing_pv else ''}"
                    )
                    return True, str(existing_file), existing_pv
        return False, None, None

    except Exception as e:
        logger.error(f"Error checking for duplicate file: {str(e)}")
        return False, None, None


def check_duplicate_entry(
    ws, pv_number: str, pdf_path: str
) -> tuple[bool, Optional[int]]:
    """
    Check if a PV number already exists in the worksheet.
    Returns a tuple of (is_duplicate: bool, row_number: Optional[int]).
    """
    if not pv_number:
        return False, None

    # Handle empty worksheet
    if ws.max_row <= 1:  # Only header row
        return False, None

    # Find the column index for "N° PV"
    pv_col_idx = None
    for idx, cell in enumerate(ws[1], 1):  # 1-based indexing for openpyxl
        if cell.value == "N° PV":
            pv_col_idx = idx
            break

    if pv_col_idx is None:
        logger.warning("Could not find 'N° PV' column in worksheet")
        return False, None

    # Check all rows in the PV number column
    try:
        for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (after header)
            cell = ws.cell(row=row_idx, column=pv_col_idx)
            cell_value = cell.value
            if (
                cell_value
                and str(cell_value).strip().upper() == str(pv_number).strip().upper()
            ):
                # Get additional context from the duplicate entry
                date_col = None
                immatri_col = None
                for idx, header in enumerate(ws[1], 1):
                    if header.value == "DATE":
                        date_col = idx
                    elif header.value == "IMMATRI":
                        immatri_col = idx
                    if date_col and immatri_col:
                        break

                # Log detailed information about the duplicate
                duplicate_info = {
                    "row": row_idx,
                    "date": ws.cell(row=row_idx, column=date_col).value
                    if date_col
                    else "Unknown",
                    "immatri": ws.cell(row=row_idx, column=immatri_col).value
                    if immatri_col
                    else "Unknown",
                }

                logger.warning(
                    f"Duplicate PV number detected!\n"
                    f"New file: {pdf_path}\n"
                    f"PV Number: {pv_number}\n"
                    f"Previous entry:\n"
                    f"  - Row: {duplicate_info['row']}\n"
                    f"  - Date: {duplicate_info['date']}\n"
                    f"  - Vehicle: {duplicate_info['immatri']}"
                )
                return True, row_idx
    except Exception as e:
        logger.error(f"Error checking for duplicate PV number: {str(e)}")
        return False, None

    return False, None


def check_previous_rejection(ws, immatri: str) -> Optional[Dict[str, Any]]:
    """
    Check if vehicle was previously rejected and within resubmission deadline.
    Returns the previous record if found and within deadline, None otherwise.
    """
    if not immatri:
        return None

    def check_worksheet(worksheet) -> Optional[Dict[str, Any]]:
        # Find column indices
        header_row = worksheet[1]
        col_indices = {}
        for idx, cell in enumerate(header_row, 1):
            col_indices[cell.value] = idx

        # Check all rows for the vehicle
        for row in worksheet.iter_rows(min_row=2):
            if row[col_indices["IMMATRI"] - 1].value == immatri:
                # Check if it was rejected
                if row[col_indices["REFUS"] - 1].value == "R":
                    # Get rejection date and resubmission deadline
                    rejection_date = datetime.strptime(
                        row[col_indices["DATE"] - 1].value, "%d/%m/%Y"
                    )

                    # Use DATE P.V as resubmission deadline
                    deadline_str = row[col_indices["DATE P.V"] - 1].value
                    if not deadline_str:
                        # If no deadline was set, assume 14 days from rejection
                        deadline_date = rejection_date + timedelta(days=14)
                        logger.warning(
                            f"No resubmission deadline found for {immatri}, "
                            f"using default 14 days from rejection date"
                        )
                    else:
                        deadline_date = datetime.strptime(deadline_str, "%d/%m/%Y")

                    today = datetime.now()

                    if today <= deadline_date:
                        # Within resubmission deadline
                        record = {
                            header.value: row[idx - 1].value
                            for header, idx in col_indices.items()
                        }
                        logger.info(
                            f"Found previous rejection for {immatri}:\n"
                            f"  Rejection date: {rejection_date.strftime('%d/%m/%Y')}\n"
                            f"  Resubmission deadline: {deadline_date.strftime('%d/%m/%Y')}\n"
                            f"  Days remaining: {(deadline_date - today).days}"
                        )
                        return record
                    else:
                        # Past resubmission deadline
                        logger.warning(
                            f"Resubmission deadline expired for {immatri}:\n"
                            f"  Rejection date: {rejection_date.strftime('%d/%m/%Y')}\n"
                            f"  Deadline was: {deadline_date.strftime('%d/%m/%Y')}\n"
                            f"  Days expired: {(today - deadline_date).days}\n"
                            f"A new full inspection is required."
                        )
                        messagebox.showwarning(
                            "Resubmission Period Expired",
                            f"Vehicle {immatri} was rejected on {rejection_date.strftime('%d/%m/%Y')}.\n"
                            f"The resubmission deadline ({deadline_date.strftime('%d/%m/%Y')}) has expired.\n"
                            f"A new full inspection is required.",
                        )
                        return None
        return None

    # First check current month's worksheet
    result = check_worksheet(ws)
    if result:
        return result

    # If not found and we're in the first few days of the month, check previous month
    today = datetime.now()
    if today.day <= 15:  # Check previous month if we're in first 15 days
        try:
            # Calculate previous month
            if today.month == 1:
                prev_month = 12
                prev_year = today.year - 1
            else:
                prev_month = today.month - 1
                prev_year = today.year

            # Get previous month's Excel file
            prev_month_name = FRENCH_MONTHS[prev_month]
            prev_excel_path = (
                config.paths["EXCEL_FOLDER"]
                / str(prev_year)
                / f"{prev_month_name} {prev_year}"
                / f"LISTING MOIS {prev_month_name} {prev_year}.xlsx"
            )

            if prev_excel_path.exists():
                logger.info(f"Checking previous month's Excel file: {prev_excel_path}")
                prev_wb = load_workbook(prev_excel_path)
                prev_ws = prev_wb.active
                result = check_worksheet(prev_ws)
                prev_wb.close()
                return result
        except Exception as e:
            logger.error(f"Error checking previous month's Excel: {str(e)}")

    return None


def get_user_input(pdf_path: str, data: Dict[str, Any]) -> Dict[str, Any]:
    """Get user input for seals number and coupon application."""
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    # Get Excel file to check for previous rejections
    excel_path, wb = get_monthly_excel()
    ws = wb.active

    # Check if this is a resubmission (automatic, no user confirmation)
    previous_data = check_previous_rejection(ws, data.get("IMMATRI"))
    if previous_data:
        if data.get("ACCEPTE") == "A":
            # Resubmission and now accepted
            data["REFUS"] = "-"
            data["C/CV"] = "C"
            data["DATE"] = datetime.now().strftime("%d/%m/%Y")
            for field in ["COUPON", "pht", "TVA", "PTTC"]:
                data[field] = 0
            # Prompt for N° SCELLES
            seals = simpledialog.askstring(
                "N° SCELLES",
                f"Enter the seals number for vehicle {data.get('IMMATRI', 'Unknown')}:",
                parent=root,
            )
            data["N° SCELLES"] = seals if seals else ""
            root.destroy()

            messagebox.showinfo(
                "Record Added",
                f"Vehicle {data['IMMATRI']} resubmission recorded as ACCEPTED with new PV.",
            )
            return data
        elif data.get("REFUS") == "R":
            # Resubmission but still refused: just record as a new refusal, normal logic
            pass
    else:
        # If not a resubmission or no previous rejection found, continue with normal flow
        if data.get("REFUS") != "R":
            seals = simpledialog.askstring(
                "N° SCELLES",
                f"Enter the seals number for vehicle {data.get('IMMATRI', 'Unknown')}:",
                parent=root,
            )
            data["N° SCELLES"] = seals if seals else ""
        else:
            data["N° SCELLES"] = ""

    # Ask about coupon
    apply_coupon = messagebox.askyesno(
        "Coupon Application",
        f"Do you want to apply a coupon for vehicle {data.get('IMMATRI', 'Unknown')}?",
    )

    if apply_coupon:
        # Calculate coupon based on PTTC and category
        pttc = int(data.get("PTTC", 0))
        category = data.get("CAT", "").strip().upper()

        # Find matching coupon rule
        coupon_amount = 0
        for price, rule in COUPON_RULES.items():
            if pttc == price and category in rule["categories"]:
                coupon_amount = rule["reduction"]
                break

        if coupon_amount > 0:
            data["COUPON"] = str(coupon_amount)
            original_pttc = int(data["PTTC"])
            data["PTTC"] = str(original_pttc - int(data["COUPON"]))

            messagebox.showinfo(
                "Coupon Applied",
                f"Coupon of {data['COUPON']} FCFA has been applied.\n"
                f"Original PTTC: {original_pttc} FCFA\n"
                f"New PTTC after coupon: {data['PTTC']} FCFA",
            )
        else:
            messagebox.showwarning(
                "No Coupon Available",
                f"No coupon available for PTTC {pttc} and category {category}",
            )
            data["COUPON"] = "0"
    else:
        data["COUPON"] = "0"

    root.destroy()
    return data


def process_pdf(pdf_path: str) -> None:
    """Process a single PDF file with enhanced error handling and monitoring."""
    start_time = time.time()
    pdf_path_obj = Path(pdf_path)
    pdf_size = pdf_path_obj.stat().st_size
    excel_path = None
    lock = None
    wb = None

    try:
        logger.info(f"Processing {pdf_path} (size: {pdf_size/1024:.2f}KB)")

        # Basic file validation
        if not pdf_path_obj.exists():
            raise ProcessingError("PDF file not found")
        if pdf_size > config.settings["MAX_PDF_SIZE"]:
            raise ProcessingError(f"PDF too large ({pdf_size/1024:.2f}KB)")

        # Calculate and log file hash
        file_hash = get_file_hash(str(pdf_path_obj))
        logger.info(f"File SHA-256: {file_hash}")

        # Extract and validate data
        data = extract_data_with_retry(str(pdf_path_obj))

        # Get user input for seals and coupon
        data = get_user_input(str(pdf_path_obj), data)

        # Process Excel file with proper locking
        excel_path, wb = get_monthly_excel()
        excel_path_obj = Path(excel_path)
        lock = FileLock(str(excel_path_obj.with_suffix('.xlsx.lock')))

        with lock:
            ws = wb.active

            # Check for duplicate entry with enhanced logging
            is_duplicate, duplicate_path, duplicate_pv = check_duplicate_file(
                str(pdf_path_obj), Path(config.paths["PDF_FOLDER"])
            )
            if is_duplicate:
                logger.warning(
                    f"Skipping processing of {pdf_path} due to duplicate PV number"
                )
                return

            # Additional duplicate check in Excel
            if "N° PV" in data:
                is_duplicate_entry, duplicate_row = check_duplicate_entry(
                    ws, data["N° PV"], str(pdf_path_obj)
                )
                if is_duplicate_entry:
                    logger.warning(
                        f"Skipping processing of {pdf_path} due to duplicate entry in Excel"
                    )
                    return

            # Check Excel row limit
            if ws.max_row >= config.settings["MAX_EXCEL_ROWS"]:
                raise ProcessingError(
                    f"Excel file has reached maximum rows ({config.settings['MAX_EXCEL_ROWS']})"
                )

            # Append data and save within the lock
            append_data(data, wb)
            try:
                wb.save(str(excel_path_obj))
                logger.info(f"Data saved into {excel_path}")
            except PermissionError:
                alternative_path = excel_path_obj.parent / f"{excel_path_obj.stem}_copy_{int(time.time())}.xlsx"
                wb.save(str(alternative_path))
                logger.warning(
                    f"Original file was locked. Data saved to: {alternative_path}"
                )
                excel_path = str(alternative_path)

        # Log processing metrics
        processing_time = time.time() - start_time
        logger.info(
            f"Processing completed successfully in {processing_time:.2f} seconds. "
            f"PDF Size: {pdf_size/1024:.2f}KB, Excel: {excel_path}"
        )

    except ProcessingError as e:
        logger.error(f"Processing error for {pdf_path}: {str(e)}")
        logger.debug(traceback.format_exc())
        raise
    except Exception as e:
        logger.error(f"Unexpected error processing {pdf_path}: {str(e)}")
        logger.debug(traceback.format_exc())
        raise
    finally:
        if lock and lock.is_locked:
            try:
                lock.release()
            except Exception as e:
                logger.error(f"Error releasing lock: {str(e)}")
        if excel_path:
            try:
                cleanup_files(excel_path)
            except Exception as e:
                logger.error(f"Error during cleanup: {str(e)}")


def process_pdfs_in_parallel(pdf_files):
    with ThreadPoolExecutor(max_workers=config.settings["MAX_WORKERS"]) as executor:
        executor.map(process_pdf, pdf_files)


class PDFProcessor:
    """Manages PDF processing queue and worker pool."""

    def __init__(self):
        self.queue = Queue()
        self.stop_event = Event()
        self.processed_count = 0
        self.error_count = 0
        self.start_time = time.time()
        self.futures = []  # Track all submitted tasks
        self.executor = None

    def add_to_queue(self, pdf_path: str) -> None:
        """Add a PDF to the processing queue."""
        self.queue.put(pdf_path)

    def process_queue(self) -> None:
        """Process PDFs in the queue using thread pool."""
        try:
            with ThreadPoolExecutor(max_workers=config.settings["MAX_WORKERS"]) as self.executor:
                while not self.stop_event.is_set():
                    try:
                        # Get PDF path from queue with timeout
                        pdf_path = self.queue.get(timeout=1)

                        # Submit processing task
                        future = self.executor.submit(self._process_single_pdf, pdf_path)
                        self.futures.append(future)

                        # Clean up completed futures
                        self._cleanup_completed_futures()

                    except queue.Empty:
                        continue
                    except Exception as e:
                        logger.error(f"Queue processing error: {str(e)}")
                        self.error_count += 1

                # Wait for all pending tasks to complete when stopping
                logger.info("Waiting for pending tasks to complete...")
                self._wait_for_pending_tasks()

        except Exception as e:
            logger.error(f"Fatal error in process_queue: {str(e)}")
            logger.debug(traceback.format_exc())
        finally:
            self._cleanup_resources()

    def _process_single_pdf(self, pdf_path: str) -> bool:
        """Process a single PDF and return success status."""
        try:
            process_pdf(pdf_path)
            self.processed_count += 1
            return True
        except Exception as e:
            logger.error(f"Failed to process {pdf_path}: {str(e)}")
            self.error_count += 1
            return False

    def _cleanup_completed_futures(self) -> None:
        """Clean up completed futures and log any errors."""
        completed = [f for f in self.futures if f.done()]
        for future in completed:
            try:
                if future.exception():
                    logger.error(f"Task failed with error: {future.exception()}")
                self.futures.remove(future)
            except Exception as e:
                logger.error(f"Error cleaning up future: {str(e)}")

    def _wait_for_pending_tasks(self) -> None:
        """Wait for all pending tasks to complete."""
        try:
            for future in self.futures:
                try:
                    future.result(timeout=30)  # Wait up to 30 seconds per task
                except TimeoutError:
                    logger.warning("Task timed out while waiting for completion")
                except Exception as e:
                    logger.error(f"Task failed while waiting for completion: {str(e)}")
        except Exception as e:
            logger.error(f"Error waiting for pending tasks: {str(e)}")

    def _cleanup_resources(self) -> None:
        """Clean up resources when shutting down."""
        try:
            if self.executor:
                self.executor.shutdown(wait=True)  # Changed to wait=True to ensure graceful shutdown
            self._log_statistics()
        except Exception as e:
            logger.error(f"Error during resource cleanup: {str(e)}")

    def _log_statistics(self) -> None:
        """Log processing statistics."""
        try:
            elapsed_time = time.time() - self.start_time
            total_processed = self.processed_count + self.error_count

            if total_processed == 0:
                logger.info("Processing Statistics:\n  No files processed yet.")
                return

            success_rate = (self.processed_count / total_processed * 100) if total_processed > 0 else 0
            avg_time = elapsed_time / total_processed if total_processed > 0 else 0

            logger.info(
                f"Processing Statistics:\n"
                f"  Total Files: {total_processed}\n"
                f"  Successful: {self.processed_count}\n"
                f"  Failed: {self.error_count}\n"
                f"  Success Rate: {success_rate:.1f}%\n"
                f"  Average Time: {avg_time:.2f}s per file\n"
                f"  Total Runtime: {elapsed_time:.2f}s"
            )
        except Exception as e:
            logger.error(f"Error logging statistics: {str(e)}")

    def stop(self) -> None:
        """Stop the processor gracefully."""
        logger.info("Stopping processor...")
        self.stop_event.set()


def get_pdf_day_folder(date_obj=None):
    """Return the correct PDF folder for a given date (default: today), creating folders if needed."""
    if date_obj is None:
        date_obj = datetime.now()
    year = date_obj.year
    month = date_obj.month
    day = date_obj.day
    month_name = FRENCH_MONTHS[month]
    folder_name = f"{month_name} {year}"
    day_folder = f"{day:02d}-{month:02d}-{year}"
    pdf_dir = config.paths["PDF_FOLDER"] / str(year) / folder_name / day_folder
    pdf_dir.mkdir(parents=True, exist_ok=True)
    return pdf_dir


def extract_control_date(pdf_path: str) -> Optional[datetime]:
    """Extract just the control date from a PDF file."""
    pdf = None
    try:
        pdf = pdfplumber.open(pdf_path)
        if not pdf.pages:
            return None

        # Extract text from first page only (date should be there)
        text = pdf.pages[0].extract_text() or ""
        lines = text.split("\n")

        # Look for control date
        for line in lines:
            if "Date du contrôle" in line and "Date of control" in line:
                parts = line.split(":")
                if len(parts) > 1:
                    date_part = parts[1].split("Catégorie")[0].strip()
                    date_match = re.search(r"(\d{2}/\d{2}/\d{4})", date_part)
                    if date_match:
                        # Convert DD/MM/YYYY to datetime object
                        date_str = date_match.group(1)
                        return datetime.strptime(date_str, "%d/%m/%Y")
        return None

    except Exception as e:
        logger.error(f"Error extracting control date from {pdf_path}: {str(e)}")
        return None
    finally:
        if pdf:
            pdf.close()


class PDFHandler(FileSystemEventHandler):
    """Handles file system events for PDFs."""

    def __init__(self, processor: PDFProcessor):
        self.processor = processor
        self._processing_paths = set()

    def on_created(self, event):
        src_path = None  # Initialize outside try block
        try:
            if event.is_directory or not event.src_path.lower().endswith(".pdf"):
                return

            src_path = Path(event.src_path)
            # Avoid duplicate processing
            if str(src_path) in self._processing_paths:
                return

            # Wait for file to be completely written
            self._wait_for_file_ready(str(src_path))

            # Move PDF to correct day folder based on control date
            # First try to get the control date from the PDF
            control_date = extract_control_date(str(src_path))

            # If we couldn't get the control date, fall back to file modification time
            if control_date is None:
                logger.warning(
                    f"Could not extract control date from {src_path}, using file modification time"
                )
                control_date = datetime.fromtimestamp(src_path.stat().st_mtime)

            # Get the target directory based on the control date
            target_dir = get_pdf_day_folder(control_date)
            target_path = target_dir / src_path.name

            # Check for duplicates in the target directory (same day's folder)
            is_duplicate, duplicate_path, duplicate_pv = check_duplicate_file(
                str(src_path), target_dir
            )
            if is_duplicate:
                # The warning log is already handled in check_duplicate_file
                return

            # Move the file if it's not already in the correct location
            if src_path.resolve() != target_path.resolve():
                # Create target directory if it doesn't exist
                target_dir.mkdir(parents=True, exist_ok=True)
                shutil.move(str(src_path), str(target_path))
                logger.info(
                    f"Moved PDF to {target_path} based on control date: {control_date.strftime('%d/%m/%Y')}"
                )
            else:
                target_path = src_path

            # Add to processing set and queue
            self._processing_paths.add(str(target_path))
            self.processor.add_to_queue(str(target_path))

        except Exception as e:
            logger.error(f"Error in file event handler: {str(e)}")
            logger.debug(traceback.format_exc())
        finally:
            # Only try to remove from processing paths if src_path was set
            if src_path is not None and str(src_path) in self._processing_paths:
                self._processing_paths.remove(str(src_path))

    def _wait_for_file_ready(self, file_path: str, timeout: int = 30) -> None:
        """Wait for file to be completely written."""
        start_time = time.time()
        last_size = -1
        path_obj = Path(file_path)

        while time.time() - start_time < timeout:
            try:
                current_size = path_obj.stat().st_size
                if current_size == last_size and current_size > 0:
                    return
                last_size = current_size
                time.sleep(0.5)
            except OSError:
                time.sleep(0.5)

        raise ProcessingError(f"Timeout waiting for file {file_path} to be ready")


if __name__ == "__main__":
    processor = None
    observer = None

    try:
        # Initialize processor
        processor = PDFProcessor()

        # Set up file watching (recursive)
        event_handler = PDFHandler(processor)
        observer = Observer()
        observer.schedule(event_handler, config.paths["PDF_FOLDER"], recursive=True)
        observer.start()

        logger.info(
            f"Started watching {config.paths['PDF_FOLDER']} for PDFs (recursive)\n"
            f"Using {config.settings['MAX_WORKERS']} workers\n"
            f"Output directory: {config.paths['EXCEL_FOLDER']}"
        )

        # Start processing queue in separate thread
        processing_thread = threading.Thread(target=processor.process_queue)
        processing_thread.start()

        try:
            while not processor.stop_event.is_set():
                time.sleep(1)
        except KeyboardInterrupt:
            logger.info("Stopping gracefully...")
            if processor:
                processor.stop()
            if observer:
                observer.stop()

        # Wait for threads to complete
        if observer:
            observer.join()
        if processing_thread:
            processing_thread.join()

        # Final statistics
        if processor:
            processor._log_statistics()

        logger.info("Application stopped cleanly")

    except Exception as e:
        logger.critical(f"Fatal error in main program: {str(e)}")
        logger.debug(traceback.format_exc())
        # Attempt cleanup
        if processor:
            try:
                processor.stop()
            except Exception as e:
                logger.error(f"Error stopping processor: {str(e)}")
                pass
        if observer:
            try:
                observer.stop()
                observer.join()
            except Exception as e:
                logger.error(f"Error stopping observer: {str(e)}")
                pass
        raise
