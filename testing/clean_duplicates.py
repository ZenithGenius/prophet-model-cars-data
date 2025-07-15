import logging
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from copy import copy
from datetime import datetime
import shutil
import argparse
from config import config

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


def copy_cell_with_format(source_cell, target_cell):
    """Copy cell value, data type, and formatting from source to target."""
    if source_cell.value is not None:
        target_cell.value = source_cell.value

    # Copy number format
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format

    # Copy data type
    if hasattr(source_cell, "data_type"):
        target_cell.data_type = source_cell.data_type

    # Copy font
    if source_cell.font:
        target_cell.font = copy(source_cell.font)

    # Copy fill
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)

    # Copy border
    if source_cell.border:
        target_cell.border = copy(source_cell.border)

    # Copy alignment
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)


def backup_excel_file(excel_path: Path) -> Path:
    """Create a backup of the Excel file before cleaning."""
    backup_path = (
        excel_path.parent
        / f"{excel_path.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    shutil.copy2(str(excel_path), str(backup_path))
    logger.info(f"Created backup at: {backup_path}")
    return backup_path


def clean_duplicates(excel_path: str) -> None:
    """Clean duplicate entries from Excel file based on PV number."""
    try:
        excel_path_obj = Path(excel_path)
        if not excel_path_obj.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        # Create backup before modifications
        backup_path = backup_excel_file(excel_path_obj)

        # Load workbook
        wb = load_workbook(excel_path)
        ws = wb.active

        # Find PV number column index
        pv_col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "N° PV":
                pv_col_idx = idx
                break

        if pv_col_idx is None:
            raise ValueError("Could not find 'N° PV' column in worksheet")

        # Dictionary to store unique entries
        unique_entries = {}
        rows_to_delete = []
        duplicate_count = 0

        # Iterate through rows (skip header)
        for row_idx in range(2, ws.max_row + 1):
            pv_number = ws.cell(row=row_idx, column=pv_col_idx).value
            if pv_number:
                pv_number = str(pv_number).strip().upper()
                if pv_number in unique_entries:
                    # This is a duplicate
                    rows_to_delete.append(row_idx)
                    duplicate_count += 1
                    logger.info(
                        f"Found duplicate PV number: {pv_number}\n"
                        f"  Original row: {unique_entries[pv_number]}\n"
                        f"  Duplicate row: {row_idx}"
                    )
                else:
                    unique_entries[pv_number] = row_idx

        if not rows_to_delete:
            logger.info("No duplicates found in the file.")
            return

        # Create new workbook with unique entries
        new_wb = Workbook()
        new_ws = new_wb.active

        # Copy column widths
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in ws.column_dimensions:
                new_ws.column_dimensions[col_letter].width = ws.column_dimensions[
                    col_letter
                ].width

        # Copy header row with formatting
        for col_idx, cell in enumerate(ws[1], 1):
            target_cell = new_ws.cell(row=1, column=col_idx)
            copy_cell_with_format(cell, target_cell)

        # Copy non-duplicate rows with formatting
        new_row_idx = 2
        for row_idx in range(2, ws.max_row + 1):
            if row_idx not in rows_to_delete:
                # Copy row height
                if row_idx in ws.row_dimensions:
                    new_ws.row_dimensions[new_row_idx].height = ws.row_dimensions[
                        row_idx
                    ].height

                for col_idx in range(1, ws.max_column + 1):
                    source_cell = ws.cell(row=row_idx, column=col_idx)
                    target_cell = new_ws.cell(row=new_row_idx, column=col_idx)
                    copy_cell_with_format(source_cell, target_cell)
                new_row_idx += 1

        # Save the cleaned file
        new_wb.save(str(excel_path))
        logger.info(
            f"Cleaning completed:\n"
            f"  Total duplicates removed: {duplicate_count}\n"
            f"  Original rows: {ws.max_row}\n"
            f"  Remaining rows: {new_ws.max_row}\n"
            f"  Backup saved at: {backup_path}"
        )

    except Exception as e:
        logger.error(f"Error cleaning duplicates: {str(e)}")
        raise


def main():
    """Main function to clean duplicates from Excel file."""
    parser = argparse.ArgumentParser(
        description="Clean duplicate entries from Excel file based on PV number."
    )
    parser.add_argument("excel_path", help="Path to the Excel file to clean")
    args = parser.parse_args()

    try:
        excel_path = Path(args.excel_path)
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        logger.info(f"Processing file: {excel_path}")
        clean_duplicates(str(excel_path))
        logger.info("Duplicate cleaning completed successfully")

    except Exception as e:
        logger.error(f"Error in main: {str(e)}")
        raise


if __name__ == "__main__":
    main()
